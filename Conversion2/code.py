
import os
import fitz
import tabula
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to extract headings and tables from a PDF
def extract_headings_and_tables(pdf_file):
    headings_and_tables = []
    pdf_document = fitz.open(pdf_file)

    for page_num in range(pdf_document.page_count):
        page = pdf_document.load_page(page_num)
        page_text = page.get_text()

        # Check if there are any tables on the page using Tabula
        tables = tabula.read_pdf(pdf_file, pages=page_num + 1, multiple_tables=True)

        if tables:
            # If tables are found, check if there is a heading before the table
            preceding_text = page_text.split('\n')[0]  # Extract the first line of text
            if preceding_text.strip():  # If not empty, consider it a heading
                headings_and_tables.append((preceding_text.strip(), tables))
        elif page_text.strip():  # If no tables are found, check if the page text is not empty
            headings_and_tables.append((page_text.strip(), None))

    return headings_and_tables

# Directory containing the PDF files
pdf_directory = "C:\Python\Conversion2\pdf_directory"

# Loop through PDF files in the directory
for pdf_file_name in os.listdir(pdf_directory):
    if pdf_file_name.endswith(".pdf"):
        # Get the full path of the PDF file
        pdf_file_path = os.path.join(pdf_directory, pdf_file_name)

        # Create a new Excel workbook
        workbook = Workbook()

        # Extract headings and tables from the PDF
        headings_and_tables = extract_headings_and_tables(pdf_file_path)

        for idx, (heading, tables) in enumerate(headings_and_tables):
            if tables:
                # If tables are found, create a new sheet for each set of heading and tables
                heading_df = pd.DataFrame([heading], columns=["Converted Succesfully"])
                tables_df = pd.concat(tables, ignore_index=True)

                # Create a new sheet with a name based on the index
                sheet_name = f"Sheet{idx + 1}"
                workbook.create_sheet(title=sheet_name)

                # Get the newly created sheet
                sheet = workbook[sheet_name]

                # Write heading and tables to the sheet
                for row in dataframe_to_rows(heading_df, index=False, header=True):
                    sheet.append(row)
                for row in dataframe_to_rows(tables_df, index=False, header=False):
                    sheet.append(row)
            else:
                # If no tables are found, create a new sheet for the heading only
                heading_df = pd.DataFrame([heading], columns=["Converted Succesfully"])

                # Create a new sheet with a name based on the index
                sheet_name = f"Sheet{idx + 1}"
                workbook.create_sheet(title=sheet_name)

                # Get the newly created sheet
                sheet = workbook[sheet_name]

                # Write heading to the sheet
                for row in dataframe_to_rows(heading_df, index=False, header=True):
                    sheet.append(row)

        # Save the workbook with a name based on the PDF file name
        output_excel_file = os.path.splitext(pdf_file_name)[0] + ".xlsx"
        workbook.save(output_excel_file)

        print(f"Processed {pdf_file_name} and saved to {output_excel_file}")



